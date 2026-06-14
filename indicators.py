import openpyxl
import pandas as pd
import numpy as np
import finplot as fplt

 
def prev_price(number, i, close_data, boolean = True):      ###اگر true بود باید از کندل های قبلی بزرگتر باشه و برعکس
    x = 0
    if boolean:
        for j in range(1, number + 1):
            if close_data[i] > close_data[i-j]:
                x += 1
        if x == number:
            return True
        return False
    else:
        for j in range(1, number + 1):
            if close_data[i] < close_data[i-j]:
                x += 1
        if x == number:
            return True
        return False

def next_price(number, i, close_data, boolean = True):      ###اگر true بود باید از کندل های بعدی بزرگتر باشه و برعکس
    x = 0
    if not boolean:
        try:
            if close_data[i+number]:
                for j in range(1, number + 1):
                    if close_data[i] < close_data[i + j]:
                        x += 1
                if x == number:
                    return True
                return False
        except IndexError:
            return False
    else:
        try:
            if close_data[i+number]:
                for j in range(1, number + 1):
                    if close_data[i] > close_data[i + j]:
                        x += 1
                if x == number:
                    return True
                return False
        except IndexError:
            return False



# def plot_volume(data_list):
#     df = pd.DataFrame(data_list, columns=['timestamp', 'volume'])
#     df['timestamp'] = pd.to_datetime(df['timestamp'])
#     df = df.set_index('timestamp')
#     lst = [
#         [0, 0, 0],  # هر آیتم باید 3 عنصر داشته باشد
#         [0, 0, 0],
#         [0, 0, 0]
#     ]
#     df2 = pd.DataFrame(lst,columns=['a','b','c'])
#     print(df2)
#     df = pd.concat([df,df2])
#     print(df)
#     fplt.create_plot('نمودار حجم معاملات', rows=1)
#     fplt.candlestick_ochl(df['a','volume','b','c'])
#     fplt.show()
#
# sample_volume_data = [
#     ['2023-01-01', 1000],
#     ['2023-01-02', 1500]
# ]
# plot_volume(sample_volume_data)


# volume = pd.DataFrame(5, index='volume')
# fplt.candlestick_ochl(volume[0, 'volume', 'volume', 0])
# fplt.show()

def indicators(name, title, smbl_data, dfr, etr_value=10, cmb_value='adjClose', res_ax = 0):    #اسم اندیکاتور اسم نماد
    data = pd.read_excel(f'download_data/{title}.xlsx')
    Date = data['date'].tolist()
    Open = data['open'].tolist()
    High = data['high'].tolist()
    Low = data['low'].tolist()
    Close = data['adjClose'].tolist()
    volume = data['volume'].tolist()
    value = data['value'].tolist()
    rates_total = len(data)
    if name == 'EMA' or name[0:4] == 'EMA-':
        print(smbl_data[dfr]['high'] + etr_value)
        return smbl_data[dfr]['high'] + etr_value
    if name == '_EMA_':
        return smbl_data[dfr]['high'] + 2
    if name == 'SMA' or name[0:4] == 'SMA-':
        data = pd.read_excel(f'download_data/{title}.xlsx')
        close = data['adjClose'].tolist()
        wcc = openpyxl.load_workbook(f'download_data/{title}.xlsx')
        row = wcc.active
        rates_total = 0
        for i in range(row.max_row):
            if row[f'B{i + 2}'].value is not None:
                rates_total += 1
        print(rates_total)
        price = []
        for i in range(rates_total):
            if i >= etr_value:
                sm = 0
                j = etr_value
                while j > 0:
                    sm += close[i - j + 1]
                    j -= 1
                average = sm / etr_value
                price.append(average)
            else:
                price.append(close[0])
        return price
    if name == '_SMA_' or name[0:6] == '_SMA_-':
        return smbl_data[dfr][cmb_value] - etr_value
        # return smbl_data[dfr]['low'] + 2
    if name == '_karimi_' or name[0:9] == '_karimi_-':
        # data = pd.read_excel(f'download_data/{title}.xlsx')
        # high = data['high'].tolist()
        # close = data['adjClose'].tolist()
        # volume = data['volume'].tolist()
        # value = data['value'].tolist()
        # rates_total = len(data)
        val = []
        price = []
        for i in range(rates_total):
            if i >= etr_value:
                sm = 0
                average = 0
                j = etr_value
                while j > 0:
                    sm += value[i-j]
                    j -= 1
                average = sm/etr_value
                price.append((value[i] / average) * High[i])
            else:
                price.append(Close[0])
        return price
    if name == '_whale_' or name[0:9] == '_whale_-':    ###اندیکاتور تشخیص حقوقی ها یا نهنگ های بازار از طریق تغییر مالکیت سربرگ حقیقی حقوقی ها
        change = pd.DataFrame(data['individual_ownership_change'])     #### دریافت ستون تغیر مالکیت به روش دیتا فریم که اگر تعداد سطر ان برابر سطر های دیتای قیمتی بود مستقیم ریترن میشود
        wc = openpyxl.load_workbook(f'download_data/{title}.xlsx')      ##خواندن فایل برای تشخیص ردیف های دیتای قیمتی که توجه شود مسیر فایل مسیری هست که فایل مین اصلی برنامه پایتون تریدر قرار دارد
        sht_close = wc.active
        total_close=0
        for k in range(1, sht_close.max_row + 1):
            if sht_close[f"{'E'}{k}"].value is not None:
                total_close += 1
        wch = openpyxl.load_workbook(f'download_data/{title}.xlsx')
        sht_change = wch.active
        total_change = 0
        for k in range(1, sht_change.max_row + 1):
            if sht_change[f"{'Y'}{k}"].value is not None:
                total_change += 1
        space_more = 0
        space_less = 0
        if total_change > total_close:      ### اگر ردیف های ستون تغیر مالکیت بیشتر از ستون قیمت بود
            space_more = total_change - total_close     ###تعین اختلاف بین سطر ها
        elif total_change < total_close:      ### برعکس بالا
            space_less = total_close - total_change
        df_zero = pd.DataFrame(0,index = range(len(Close)), columns=['zc'])   ###یک دیتا فریم با مقدار 0 و به تعداد کندل های قیمتی میسازد چون ایندکس اش رو تعداد داده های کلوز گذاشتیم
        fplt.plot(df_zero,ax=res_ax, color='red')       ## در اینجا دیتا فریم قبلی و ریس ای ایکس که همان ای ایکس نمودار اصلی قیمتی هست و در متد چارت فرساده شده را گذاشتیم
        if space_more:
            change= change.loc[space_more:]   ###به تعداد اسپیس مور یا همان اختلافشان از ابتدای دیتافریم جدا کن و مابقی را جایگذاری کن
            return change.reset_index(drop=True)    ### چون x خانه از دیتا فریم را در خط قبلی پاک کردیم و ایندکس ما x خانه جابجا شده باید ترتیب ایندکس ها را با ریست ایندکس از 0 شروع کنیم
        elif space_less:
            change=change.dropna()      #### خانه های نال را پاک میکند
            df_con = pd.DataFrame(0, index=range(space_less), columns=change.columns)   ###به تعداد خانه هایی که از دیتای قیمتی یا همان تعداد کندال ها کم دارد یک دیتا فریم با مقدار 0 میسازد که که تعداد ایندکس ها یا سطر های ان به اندازه سطر هایی است که کم دارد یا همان space less و ستون یا ستون های ان که در اینحا یکی است باید همنام ستون دیتای قیمتی باشد
            change = pd.concat([df_con, change])    ###دیتا فریم قبلی که مقادیر 0 دارد را به ابتدای دیتافریم change که همان ستون درصد تغیر مالکیت فایل اکسل هست اضافه میکند
            return change.reset_index(drop=True)
        return change
    if name == '_volume_':
        return volume
        # volume = pd.DataFrame(data['volume'])
        # fplt.candlestick_ochl(volume[0, 'volume', 'volume', 0], ax=res_ax)
    if name == 'candle-pattern':
        sma = []
        dframe = pd.DataFrame(data)
        dframe['Date'] = pd.to_datetime(dframe['date'])
        dframe['High'] = pd.to_numeric(dframe['high'])
        for i, (dte, h) in enumerate(zip(dframe['Date'], dframe['High'])):
            body = abs(Close[i] - Open[i])
            prev_body = abs(Close[i - 1] - Open[i - 1])
            two_body = abs(Close[i - 2] - Open[i - 2])
            three_body = abs(Close[i - 3] - Open[i - 3])
            four_body = abs(Close[i - 4] - Open[i - 4])

            upper_shadow = High[i] - max(Close[i], Open[i])
            prev_upper = High[i - 1] - max(Close[i - 1], Open[i-1])
            two_upper = High[i - 2] - max(Close[i - 2], Open[i - 2])
            three_upper = High[i - 3] - max(Close[i - 3], Open[i - 3])
            four_upper = High[i - 4] - max(Close[i - 4], Open[i - 4])

            lower_shadow = min(Close[i], Open[i]) - Low[i]
            prev_lower = min(Close[i-1], Open[i-1]) - Low[i-1]
            two_lower = min(Close[i-2], Open[i-2]) - Low[i-2]
            three_lower = min(Close[i-3], Open[i-3]) - Low[i-3]
            four_lower = min(Close[i-4], Open[i-4]) - Low[i-4]

            min_price = min(Close[i], Open[i])
            prev_min = min(Close[i - 1], Open[i - 1])
            two_min = min(Close[i - 2], Open[i - 2])
            three_min = min(Close[i - 3], Open[i - 3])
            four_min = min(Close[i - 4], Open[i - 4])

            max_price = max(Close[i], Open[i])
            prev_max = max(Close[i - 1], Open[i - 1])
            two_max = max(Close[i - 2], Open[i - 2])
            three_max = max(Close[i - 3], Open[i - 3])
            four_max = max(Close[i - 4], Open[i - 4])

            bullish = (Open[i] < Close[i])
            prev_bullish = (Close[i - 1] > Open[i - 1])
            two_bullish = (Close[i - 2] > Open[i - 2])
            three_bullish = (Close[i - 3] > Open[i - 3])
            four_bullish = (Close[i - 4] > Open[i - 4])

            if (body <= (upper_shadow + lower_shadow) * 0.3 and upper_shadow > body and lower_shadow > body):
                fplt.add_text((dte, h+High[i] / 40), 'Doubt')
            if Open[i] == Close[i] and Open[i] == Low[i] and High[i] >= Open[i] + Close[i] * 0.02:
                fplt.add_text((dte, h+High[i] / 40), '--Gravestone')
            if Open[i] == Close[i] and Open[i] == High[i] and Low[i] <= Open[i] - Close[i] * 0.02:
                fplt.add_text((dte, h+High[i] / 40), '++Dragonfly')

            if i >= 10:
                sm = 0
                j = 10
                while j > 0:
                    sm += Close[i - j + 1]
                    j -= 1
                average = sm / 10
                sma.append(average)
                ###### الگو های بازگشتی و ادامه دهنده
                ### روند صعودی
                if sma[i] > sma[i - 10]:
                    # if i + 2 <= rates_total:
                    ### 1-candle
                    if (upper_shadow < lower_shadow / 5 and body > upper_shadow and body < lower_shadow / 2):
                        fplt.add_text((dte, h + High[i] / 40), '-Hangingman')
                    if (Open[i] == High[i] and Open[i] > Open[i - 1] and Open[i] > Open[i - 2] and Open[i] > Close[i]
                            and Open[i - 1] < Close[i - 1] and Open[i - 2] < Close[i - 2] and
                            Close[i] < (Open[i - 1] + Close[i - 2]) / 2 and lower_shadow < body / 10):
                        fplt.add_text((dte, h + High[i] / 40), '-Belthold')
                    ### 2-candle
                    if (prev_body > prev_upper + prev_lower and Close[i-1] > Open[i-1] and Open[i] > Close[i-1] and Close[i] < Open[i]
                            and Close[i] > Open[i-1] and Close[i] <= (Open[i-1] + Close[i-1]) / 2 and High[i] > High[i-1]):
                        fplt.add_text((dte, h + High[i] / 40), '-Darkcloud')
                    if (Open[i - 1] < Close[i - 1] and prev_body > prev_lower + prev_upper and body < prev_body and
                            Open[i] < Close[i - 1] and Close[i] <= Open[i] and Close[i] > Open[i - 1]):
                        fplt.add_text((dte, h + High[i] / 40), '-Harami')
                    if (Open[i - 1] < Close[i - 1] and Open[i] > Close[i - 1] and Close[i] < Open[i - 1] and
                        prev_body>=prev_lower+prev_upper):
                        fplt.add_text((dte, h + High[i] / 40), '-Engulfing')
                    if(Open[i-1]<Close[i-1] and prev_body > prev_upper + prev_lower and Open[i] == Close[i] and
                        upper_shadow == lower_shadow and Open[i] > Close[i-1]) and upper_shadow > 0:
                        fplt.add_text((dte, h + High[i] / 40), '-Dojistar')
                    if (prev_upper + prev_lower < prev_body and Open[i - 1] < Close[i - 1] and body > upper_shadow + lower_shadow and
                            Open[i] > Close[i - 1] and Close[i] == Close[i - 1] and Open[i] > Close[i]):
                        fplt.add_text((dte, h + High[i] / 40), '-Meetingline')
                    if (Open[i - 1] < Close[i - 1] and prev_lower == 0 and prev_upper == 0 and Open[i] > Close[i] and Open[i] < Open[i - 1]
                        and upper_shadow == 0 and lower_shadow == 0):
                        fplt.add_text((dte, h + High[i] / 40), '-kicking')
                    if (Open[i - 1] > Close[i - 1] and Open[i] < Close[i] and Open[i] > Open[i - 1] and prev_lower == 0 and prev_upper == 0
                            and upper_shadow == 0 and lower_shadow == 0):
                        fplt.add_text((dte, h + High[i] / 40), '+=kicking')
                    if (prev_bullish and bullish and prev_body > prev_lower + prev_upper and body and max_price == prev_max and
                        Low[i] > prev_min and min_price >= (prev_max + prev_min) / 2):
                        fplt.add_text((dte, h + High[i] / 40), '-Matchinghigh')
                    if (not prev_bullish and bullish and prev_body > prev_upper + prev_lower and body > upper_shadow + lower_shadow and min_price >= prev_max):
                        fplt.add_text((dte, h + High[i] / 40), '+=Separatinglines')
                    if (prev_bullish and prev_body > prev_upper + prev_lower and not bullish and body < prev_body and min_price >= prev_max and lower_shadow == 0 and body):
                        fplt.add_text((dte, h + High[i] / 40), '+=Onneckline')
                    ### 3-candle
                    if(two_body > two_lower + two_upper and Close[i-2] > Open[i-2] and Close[i] < Open[i] and
                        Open[i-1] > Close[i-2] and Close[i-1] > Close[i-2] and prev_body < two_body / 2 and prev_body * 2 < body and
                        Open[i] <= Close[i-1] and Close[i] <= (Open[i-2] + Close[i-2])/2 and Close[i] > Open[i-2]):
                        fplt.add_text((dte, h + High[i] / 40), '-Eveningstar')

                    # if(prev_lower < prev_upper / 5 and prev_body > prev_lower and prev_body < prev_upper / 2
                    #     and High[i] < High[i-1] and Close[i]<Close[i-1] and Open[i]>Close[i] and body > prev_body and Open[i-2] < Close[i-2]):
                    #     fplt.add_text((dte, h + High[i] / 40), '-Shootingstar')
                    if (prev_lower == 0 and prev_body and prev_body < prev_upper / 2 and not bullish and High[i] < prev_min and body > prev_body):
                        fplt.add_text((dte, h + High[i] / 40), '-Shootingstar')
                    if (Open[i - 2] < Close[i - 2] and two_body > two_lower + two_upper and Open[i-1] == Close[i - 1] and
                       Open[i - 1] > Close[i - 2] and Open[i] > Close[i] and Open[i-1]>Open[i] and body>upper_shadow+lower_shadow):
                        fplt.add_text((dte, h + High[i] / 40), '-Abandonedbaby')
                    if (two_body == 0 and prev_body == 0 and body == 0 and Open[i - 1] > Open[i - 2] and Open[i] < Open[i - 1]):
                        fplt.add_text((dte, h + High[i] / 40), '-Tristar')
                    if (Open[i - 2] > Close[i - 2] and Open[i - 1] < Open[i - 2] and Open[i - 1] > Close[i - 2] and
                            Close[i - 1] < Close[i - 2] and
                            Open[i] < Open[i - 1] and Open[i] > Close[i - 1] and Close[i] < Close[i - 1]):
                        fplt.add_text((dte, h + High[i] / 40), '-Threeblackcrows')
                    if (two_bullish and two_body > two_lower + two_upper and Open[i - 1] >= Close[i - 1] and prev_max <= two_max and
                        prev_min > two_min and not bullish and max_price < prev_max and min_price < prev_min):
                        fplt.add_text((dte, h + High[i] / 40), '-Threeinsidedown')
                    if (two_bullish and not prev_bullish and not bullish and prev_max > two_max and prev_min < two_min and
                        min_price < prev_min and max_price < prev_max and two_body < prev_body):
                        fplt.add_text((dte, h + High[i] / 40), '-Threeoutsidedown')
                    if (not two_bullish and not prev_bullish and not bullish and two_body > two_upper + two_lower and prev_body > prev_upper + prev_lower
                        and body > upper_shadow + lower_shadow and prev_max == two_min and max_price == prev_min):
                        fplt.add_text((dte, h + High[i] / 40), '-Identicalthreecrows')
                    if (two_bullish and prev_bullish and Open[i] <= Close[i] and two_body > two_upper + two_lower and prev_body > prev_upper + prev_lower
                            and prev_max > two_max and prev_min > two_min and min_price > prev_max):
                        fplt.add_text((dte, h + High[i] / 40), '-Deliberation')
                    if (two_bullish and two_body > two_lower + two_upper and not prev_bullish and prev_body < two_body and prev_min > two_max and
                            not bullish and min_price > two_max and min_price < prev_min and max_price > prev_max):
                        fplt.add_text((dte, h + High[i] / 40), '-Upsidegaptwocrows')
                    if (two_bullish and two_body > two_lower + two_upper and not prev_bullish and prev_body < two_body and prev_min > two_max and
                            not bullish and min_price < two_max and max_price >= prev_min):
                        fplt.add_text((dte, h + High[i] / 40), '-Twocrows')
                    if (two_bullish and prev_bullish and bullish and two_body > two_lower + two_upper and prev_body < two_body and body < prev_body and
                        prev_lower + prev_upper > two_lower + two_upper and upper_shadow + lower_shadow > prev_lower + prev_upper and prev_max > two_max and
                        prev_min > two_min and min_price > prev_min and max_price > prev_max):
                        fplt.add_text((dte, h + High[i] / 40), '-Advanceblock')
                    if (two_bullish and prev_bullish and not bullish and two_body > two_lower + two_upper and prev_body > prev_lower + prev_upper
                        and two_max < prev_min and max_price < prev_max and max_price > prev_min and min_price < prev_min and min_price >= (two_min + two_max) / 2):
                        fplt.add_text((dte, h + High[i] / 40), '+=Upsidetasukigap')
                    if (two_bullish and prev_bullish and bullish and two_body > two_lower + two_upper and prev_min > two_max and prev_body < two_body
                        and body < prev_body and min_price > two_max and max_price < prev_max):
                        fplt.add_text((dte, h + High[i] / 40), '+=Sidebysidewhitelines')
                    ### 4-candle
                    if (three_bullish and two_bullish and prev_bullish and not bullish and three_body > three_lower + three_upper and two_body>two_lower+two_upper
                        and prev_body > prev_upper + prev_lower and body > upper_shadow + lower_shadow and three_max < two_max and two_max < prev_max
                        and prev_max < max_price and min_price < three_min and min_price < two_min and min_price < prev_min):
                        fplt.add_text((dte, h + High[i] / 40), '-Threelinestrike')
                    ### 5-candle
                    if(four_body > four_lower + four_upper and four_body > three_body and four_body > two_body and
                         four_body > prev_body and Close[i-4] > Open[i-4] and Close[i-3] < Open[i-3] and
                         Close[i-2] < Open[i-2] and Close[i-1] < Open[i-1] and Close[i] > Open[i] and Close[i] > Close[i-4] and
                         Open[i] <= (Open[i-4] + Close[i-4])/2 and Open[i-3] <= High[i-4] and Close[i-3] > Open[i-4] and
                         Open[i-2] <= Close[i-4] and Close[i-2] > Open[i-4] and Open[i-1] <= Close[i-4] and Close[i-1] > Low[i-4]):
                        fplt.add_text((dte, h + High[i] / 40), '+=Risingthree')
                    if (four_bullish and four_body > four_lower + four_upper and three_bullish and Open[i - 3] > Close[i - 4] and
                        two_min > Open[i - 3] and two_max > Close[i - 3] and prev_max > two_max and prev_min > two_min and
                        Open[i] < prev_max and Close[i] < three_min and Open[i] > prev_min):
                        fplt.add_text((dte, h + High[i] / 40), '-Breakaway')
                    if (four_bullish and three_bullish and two_bullish and prev_bullish and not bullish and three_max > four_max
                        and three_min > four_min and two_min > three_min and two_max > three_max and prev_min > two_min and prev_max > two_max
                        and four_body > four_upper + four_lower and three_body > three_lower + three_upper and two_body > two_lower + two_upper and prev_upper == 0 and
                        prev_lower >= prev_body and prev_body and max_price < prev_min and min_price < two_min and body > upper_shadow + lower_shadow):
                        fplt.add_text((dte, h + High[i] / 40), '-Laddertop')

                ###روند نزولی
                else:
                    ### 1-candle
                    if (upper_shadow < lower_shadow / 5 and body > upper_shadow and body < lower_shadow / 2):
                        fplt.add_text((dte, h + High[i] / 40), '+Hammer')
                    if(Open[i] == Low[i] and Open[i] < Open[i-1] and Open[i]<Open[i-2] and Open[i]<Close[i]
                        and Open[i-1]>Close[i-1] and Open[i-2]>Close[i-2] and Close[i] > (Open[i-1] + Close[i-2])/2
                        and upper_shadow < body/10):
                        fplt.add_text((dte, h + High[i] / 40), '+Belthold')
                    ### 2-candle
                    if(Close[i-1] < Open[i-1] and Open[i] < Close[i-1] and Close[i] > Open[i]
                            and Close[i] < Open[i-1] and Close[i] >= (Open[i-1] + Close[i-1]) / 2):
                        fplt.add_text((dte, h + High[i] / 40), '+Piercing')
                    if(Open[i-1] > Close[i-1] and prev_body > prev_lower+prev_upper and body < prev_body and
                        Open[i]>Close[i-1] and Close[i]>=Open[i] and Close[i] < Open[i-1]):
                        fplt.add_text((dte, h + High[i] / 40), '+Harami')
                    if (Open[i - 1] > Close[i - 1] and Open[i] < Close[i - 1] and Close[i] > Open[i - 1] and
                        prev_body>=prev_lower+prev_upper):
                        fplt.add_text((dte, h + High[i] / 40), '+Engulfing')
                    if (Open[i - 1] > Close[i - 1] and prev_body > prev_upper + prev_lower and Open[i] == Close[i] and
                            upper_shadow == lower_shadow and Open[i] < Close[i - 1] and upper_shadow > 0):
                        fplt.add_text((dte, h + High[i] / 40), '+Dojistar')
                    if (prev_body > prev_upper + prev_lower and Open[i - 1] > Close[i - 1] and body > upper_shadow + lower_shadow and
                            Open[i] < Close[i - 1] and Close[i] == Close[i - 1] and Open[i] < Close[i]):
                        fplt.add_text((dte, h + High[i] / 40), '+Meetingline')
                    if (Open[i - 1] > Close[i - 1] and Open[i] < Close[i] and Open[i] > Open[i - 1] and prev_lower == 0 and prev_upper == 0
                            and upper_shadow == 0 and lower_shadow == 0):
                        fplt.add_text((dte, h + High[i] / 40), '+kicking')
                    if (Open[i - 1] < Close[i - 1] and prev_lower == 0 and prev_upper == 0 and Open[i] > Close[i] and Open[i] < Open[i - 1]
                            and upper_shadow == 0 and lower_shadow == 0):
                        fplt.add_text((dte, h + High[i] / 40), '-=kicking')
                    if (not prev_bullish and prev_body > prev_upper + prev_lower and not bullish and min_price == prev_min and
                        High[i] < prev_max and max_price <= (prev_max + prev_min) / 1.8 and body):
                        fplt.add_text((dte, h + High[i] / 40), '+Matchinglow')
                    if (not prev_bullish and not bullish and prev_body > prev_upper + prev_lower and prev_max > High[i] and prev_min < Low[i] and body):
                        fplt.add_text((dte, h + High[i] / 40), '+Homingpigeon')
                    if (prev_bullish and not bullish and prev_body > prev_upper + prev_lower and body > upper_shadow + lower_shadow and max_price <= prev_min):
                        fplt.add_text((dte, h + High[i] / 40), '-=Separatinglines ')
                    if (not prev_bullish and prev_body > prev_upper + prev_lower and bullish and body < prev_body and max_price <= prev_min and upper_shadow == 0 and body):
                        fplt.add_text((dte, h + High[i] / 40), '-=Onneckline')
                    ### 3-candle
                    if(Close[i-2] < Open[i-2] and Close[i] > Open[i] and Open[i-1] < Close[i-2] and
                        Close[i-1] < Close[i-2] and prev_body < two_body / 2 and prev_body * 2 < body and
                        Open[i] >= Close[i-1] and Close[i] >= (Open[i-2] + Close[i-2])/2 and Close[i] < Open[i-2]):
                        fplt.add_text((dte, h + High[i] / 40), '+Morningstar')
                    # if(prev_lower < prev_upper / 5 and prev_body > prev_lower and prev_body < prev_upper / 2
                    #     and Open[i] > Open[i-1] and Open[i] < Close[i] and body > prev_body and Open[i-2] > Close[i-2]):
                    #     fplt.add_text((dte, h + High[i] / 40), '+Invertedhammer')
                    if (prev_lower == 0 and prev_body and prev_body < prev_upper / 2 and bullish and Low[i] > prev_max and body > prev_body):
                        fplt.add_text((dte, h + High[i] / 40), '+Invertedhammer')
                    if(Open[i-2] > Close[i-2] and two_body > two_lower + two_upper and Open[i-1] == Close[i-1] and
                        Open[i-1] < Close[i-2] and Open[i] < Close[i] and Open[i-1] < Open[i] and body > upper_shadow + lower_shadow):
                        fplt.add_text((dte, h + High[i] / 40), '+Abandonedbaby')
                    if (two_body == 0 and prev_body == 0 and body == 0 and Open[i - 1] < Open[i - 2] and Open[i] > Open[i - 1]):
                        fplt.add_text((dte, h + High[i] / 40), '+Tristar')
                    if (Open[i - 2] < Close[i - 2] and Open[i - 1] > Open[i-2] and Open[i-1] < Close[i - 2] and
                       Close[i - 1] > Close[i - 2] and Open[i] > Open[i-1] and Open[i] < Close[i-1] and Close[i] > Close[i-1]):
                        fplt.add_text((dte, h + High[i] / 40), '+Threewhitesoldier')
                    if (not two_bullish and two_body > two_lower + two_upper and Open[i - 1] <= Close[i - 1] and prev_max <= two_max and
                        prev_min > two_min and bullish and max_price > prev_max and min_price <= prev_max and min_price >= prev_min):
                        fplt.add_text((dte, h + High[i] / 40), '+Threeinsideup')
                    if (not two_bullish and prev_bullish and bullish and prev_max > two_max and prev_min < two_min and
                            min_price > prev_min and max_price > prev_max and two_body < prev_body):
                        fplt.add_text((dte, h + High[i] / 40), '+Threeoutsideup')
                    if (not two_bullish and two_body > two_lower + two_upper and not prev_bullish and prev_upper == 0 and prev_max < two_max
                        and prev_min > two_min and Low[i - 1] < Low[i - 2] and bullish and max_price <= prev_min and min_price >= Low[i - 1]):
                        fplt.add_text((dte, h + High[i] / 40), '+Uniquethreeriver')
                    if (not two_bullish and two_upper == 0 and two_lower > two_body / 3 and not prev_bullish and two_body > prev_body and
                        two_max > prev_upper and prev_max < two_max and prev_min <= two_min and not bullish and body < two_body and
                        upper_shadow == 0 and lower_shadow == 0 and min_price <= prev_min and max_price < prev_max and body):
                        fplt.add_text((dte, h + High[i] / 40), '+Threestarsinthesouth')
                    if (not two_bullish and two_lower == 0 and two_body > two_upper and prev_bullish and prev_min >= (two_min + two_max) / 2
                            and prev_max > High[i - 2] and not bullish and max_price > prev_max and min_price == two_min and prev_body > prev_upper + prev_lower
                            and body > upper_shadow and lower_shadow == 0):
                        fplt.add_text((dte, h + High[i] / 40), '+Sticksandwich')
                    if (not two_bullish and not prev_bullish and bullish and two_body > two_lower + two_upper and prev_body > prev_lower + prev_upper
                        and two_min > prev_max and min_price < prev_max and min_price > prev_min and max_price > prev_max and max_price <= (two_min + two_max) / 2):
                        fplt.add_text((dte, h + High[i] / 40), '-=Downsidetasukigap')
                    if (not two_bullish and prev_bullish and bullish and two_body > two_lower + two_upper and prev_max < two_min and prev_body < two_body
                        and body < prev_body and max_price < two_min and min_price > prev_min):
                        fplt.add_text((dte, h + High[i] / 40), '-=Sidebysidewhitelines')
                    ### 4-candle
                    if (not three_bullish and three_body and three_upper == 0 and three_lower == 0 and not two_bullish and
                        three_body >= two_body and two_body and two_upper == 0 and two_lower == 0 and two_max < three_max and two_min < three_min
                        and not prev_bullish and prev_body < two_body and prev_min < two_min and prev_max and two_max and
                        High[i - 1] > two_min and not bullish and max_price >= High[i - 1] and min_price < prev_min and prev_body
                        and body > three_body * 0.75 and body <= three_body * 1.3):
                        fplt.add_text((dte, h + High[i] / 40), '+Concealingbabyswallow')
                    if (not three_bullish and not two_bullish and not prev_bullish and bullish and three_body > three_lower + three_upper and two_body > two_lower + two_upper
                        and prev_body > prev_upper + prev_lower and body > upper_shadow + lower_shadow and three_min > two_min and two_min > prev_min
                        and prev_min > min_price and max_price > three_max and max_price > two_max and max_price > prev_max):
                        fplt.add_text((dte, h + High[i] / 40), '+Threelinestrike')
                    ### 5-candle
                    if (four_body > four_lower + four_upper and four_body > three_body and four_body > two_body and
                        four_body > prev_body and Close[i - 4] < Open[i - 4] and Close[i - 3] > Open[i - 3] and
                        Close[i-2] > Open[i-2] and Close[i - 1] > Open[i - 1] and Close[i] < Open[i] and Close[i] < Close[i - 4] and
                        Open[i] >= (Open[i-4] + Close[i - 4]) / 2 and Open[i - 3] >= Low[i - 4] and Close[i - 3] < Open[i - 4] and
                        Open[i-2] >= Close[i-4] and Close[i-2] < Open[i-4] and Open[i-1] > Close[i-4] and Close[i-1] <= High[i-4]):
                        fplt.add_text((dte, h + High[i] / 40), '-=Fallingthree')
                    if (not four_bullish and four_body > four_lower + four_upper and not three_bullish and Open[i - 3] < Close[i - 4] and
                        two_max < Open[i - 3] and two_min < Close[i - 3] and prev_max < two_max and prev_min < two_min and
                        Open[i] < prev_max and Close[i] > three_max and Open[i] > prev_min):
                        fplt.add_text((dte, h + High[i] / 40), '+Breakaway')
                    if (not four_bullish and not three_bullish and not two_bullish and not prev_bullish and bullish and three_max < four_max
                        and three_min < four_min and two_min < three_min and two_max < three_max and prev_min < two_min and prev_max < two_max
                        and four_body > four_upper + four_lower and three_body > three_lower + three_upper and two_body > two_lower + two_upper and prev_lower == 0 and
                        prev_upper >= prev_body and prev_body and min_price > prev_max and max_price > two_max and body>upper_shadow+lower_shadow):
                        fplt.add_text((dte, h + High[i] / 40), '+Ladderbottom')

            else:
                sma.append(Close[i])
    if name == 'chart-pattern':
        sma = []
        dframe = pd.DataFrame(data)
        dframe['Date'] = pd.to_datetime(dframe['date'])
        dframe['Low'] = pd.to_numeric(dframe['low'])
        ### double top & bottom
        trend = 315
        number = 105
        neck_number = 63
        mid_number = 126
        percentage = 0.18
        ### double top
        double_top = 0
        holder_i = 0
        holder_i2 = 0
        neck_price = 0
        neck_i = 0
        arr_holder_i = []
        arr_double_top = []
        volume_double_top = 0
        # init = False
        ### double bottom
        double_bottom = 0
        bottom_holder_i = 0
        bottom_holder_i2 = 0
        bottom_neck_price = 0
        bottom_neck_i = 0
        bottom_arr_holder_i = []
        bottom_arr_double_top = []
        volume_double_bottom = 0
        # bottom_init = False
        ### head and shoulders top


        #### آرایه  برای سطوح حمایت و مقاومت
        # resistance = []
        # support = []
        for k in range(1, 301):
            sma = []
            for i, (dte, l) in enumerate(zip(dframe['Date'], dframe['Low'])):
                if i >= trend:
                    sm = 0
                    j = trend
                    while j > 0:
                        sm += Close[i - j + 1]
                        j -= 1
                    average = sm / trend
                    sma.append(average)
                    ###### الگو های بازگشتی و ادامه دهنده
                    ### روند صعودی
                    if sma[i] > sma[i - trend]:
                        #### double_top
                        if prev_price(number, i, Close) and double_top == 0 and next_price(neck_number, i, Close):    ### تشخیص سقف اول
                            double_top = 1
                            holder_i = i
                        elif (double_top == 1 and prev_price(neck_number, i, Close) and next_price(number, i, Close) and
                            Close[i] <= Close[holder_i] + Close[holder_i] * percentage and Close[i] >= Close[holder_i] - Close[holder_i] * percentage):     ###سقف دوم اگر  فاصله اش با سقف قبلی کمتر از میدنامبر باشه و کمتر 3درصد سقف اول یا بیشتر 3 درصد سقف دوم باشه
                            double_top = 2
                            holder_i2 = i
                            volume_double_top = volume[i]
                            neck_price = Close[holder_i]
                            for j in range(holder_i, i):       ### پیدا کردن دره
                                if Close[j] < neck_price:
                                    neck_price = Close[j]
                                    neck_i = j
                        elif double_top == 2 and Close[i] <= neck_price - neck_price * 0.03 and volume_double_top <= volume[i]:  ### قطع گردن
                            double_top = 3
                        if double_top == 1 and i > holder_i + mid_number:
                            double_top = 0
                        if double_top == 2 and i > holder_i2 + mid_number:
                            double_top = 0
                        if double_top == 3:
                            double_top = 0
                            bifour = True
                            for m, n in zip(arr_holder_i, arr_double_top):
                                if m <= i <= n:
                                    bifour = False
                            if bifour:
                                arr_double_top.append(i)
                                arr_holder_i.append(holder_i)
                                fplt.add_text((dframe['Date'][holder_i], Close[holder_i]), f'---R---: {k}',color='red')
                                fplt.add_text((dframe['Date'][holder_i2], Close[holder_i2]), f'---R2---: {k}',color='red')
                                fplt.add_text((dframe['Date'][neck_i], Close[neck_i]), f'---S---: {k}',color='red')
                                target = Close[holder_i] - Close[neck_i]
                                target -= Close[neck_i]
                                fplt.add_text((dte, l - Low[i] / 40), f'-Doubletop: {target}')
                        ### double bottom
                        if double_bottom == 2 and Close[i] >= bottom_neck_price + bottom_neck_price * 0.03 and volume_double_bottom <= volume[i]:  ### قطع گردن
                            double_bottom = 3
                        if double_bottom == 1 and i > bottom_holder_i + mid_number:
                            double_bottom = 0
                        if double_bottom == 2 and i > bottom_holder_i2 + mid_number:
                            double_bottom = 0
                        if double_bottom == 3:
                            double_bottom = 0
                            bifour = True
                            for m, n in zip(bottom_arr_holder_i, bottom_arr_double_top):
                                if m <= i <= n:
                                    bifour = False
                            if bifour:
                                bottom_arr_double_top.append(i)
                                bottom_arr_holder_i.append(bottom_holder_i)
                                fplt.add_text((dframe['Date'][bottom_holder_i], Close[bottom_holder_i]),f'---R---: {k}', color='orange')
                                fplt.add_text((dframe['Date'][bottom_holder_i2], Close[bottom_holder_i2]),f'---R2---: {k}', color='orange')
                                fplt.add_text((dframe['Date'][bottom_neck_i], Close[bottom_neck_i]), f'---S---: {k}',color='orange')
                                target = abs(Close[bottom_neck_i] - Close[bottom_holder_i])
                                target += Close[bottom_neck_i]
                                fplt.add_text((dte, l - Low[i] / 40), f'+Doublebottom: {target}')

                    ### روند نزولی
                    else:
                        ### double top
                        if double_top == 2 and Close[i] <= neck_price - neck_price * 0.03 and volume_double_top <= volume[i]:  ### قطع گردن
                            double_top = 3
                        if double_top == 1 and i > holder_i + mid_number:
                            double_top = 0
                        if double_top == 2 and i > holder_i2 + mid_number:
                            double_top = 0
                        if double_top == 3:
                            double_top = 0
                            ### برای رسم سطوح حمایت و مقاومت
                            # resistance.append(three_holder_i)
                            # support.append(neck_i)
                            bifour = True
                            for m, n in zip(arr_holder_i, arr_double_top):
                                if m <= i <= n:
                                    bifour = False
                            if bifour:
                                arr_double_top.append(i)
                                arr_holder_i.append(holder_i)
                                fplt.add_text((dframe['Date'][holder_i], Close[holder_i]), f'---R---: {k}',color='red')
                                fplt.add_text((dframe['Date'][holder_i2], Close[holder_i2]), f'---R2---: {k}',color='red')
                                fplt.add_text((dframe['Date'][neck_i], Close[neck_i]), f'---S---: {k}',color='red')
                                target = Close[holder_i] - Close[neck_i]
                                target -= Close[neck_i]
                                fplt.add_text((dte, l - Low[i] / 40), f'-Doubletop: {target}')
                    ### double bottom
                    if prev_price(number, i, Close, False) and double_bottom == 0 and next_price(neck_number, i, Close,False):  ### تشخیص سقف اول
                        double_bottom = 1
                        bottom_holder_i = i
                    elif (double_bottom == 1 and prev_price(neck_number, i, Close,False) and next_price(number, i, Close,False) and
                        Close[i] <= Close[bottom_holder_i] + Close[bottom_holder_i] * percentage and Close[i] >= Close[bottom_holder_i] -
                        Close[bottom_holder_i] * percentage):  ###سقف دوم اگر  فاصله اش با سقف قبلی کمتر از میدنامبر باشه و کمتر 3درصد سقف اول یا بیشتر 3 درصد سقف دوم باشه
                        double_bottom = 2
                        bottom_holder_i2 = i
                        bottom_neck_price = Close[bottom_holder_i]
                        volume_double_bottom = volume[i]
                        for j in range(bottom_holder_i, i):  ### پیدا کردن دره
                            if Close[j] > bottom_neck_price:
                                bottom_neck_price = Close[j]
                                bottom_neck_i = j
                    elif double_bottom == 2 and Close[i] >= bottom_neck_price + bottom_neck_price * 0.03 and volume_double_bottom <= volume[i]:  ### قطع گردن
                        double_bottom = 3
                    if double_bottom == 1 and i > bottom_holder_i + mid_number:
                        double_bottom = 0
                    if double_bottom == 2 and i > bottom_holder_i2 + mid_number:
                        double_bottom = 0
                    if double_bottom == 3:
                        double_bottom = 0
                        bifour = True
                        for m, n in zip(bottom_arr_holder_i, bottom_arr_double_top):
                            if m <= i <= n:
                                bifour = False
                        if bifour:
                            bottom_arr_double_top.append(i)
                            bottom_arr_holder_i.append(bottom_holder_i)
                            fplt.add_text((dframe['Date'][bottom_holder_i], Close[bottom_holder_i]), f'---R---: {k}', color='orange')
                            fplt.add_text((dframe['Date'][bottom_holder_i2], Close[bottom_holder_i2]), f'---R2---: {k}', color='orange')
                            fplt.add_text((dframe['Date'][bottom_neck_i], Close[bottom_neck_i]), f'---S---: {k}', color='orange')
                            target = abs(Close[bottom_neck_i] - Close[bottom_holder_i])
                            target += Close[bottom_neck_i]
                            fplt.add_text((dte, l - Low[i] / 40), f'+Doublebottom: {target}')
                            # fplt.add_text((dte, target), f'****: {k}')
                else:
                    sma.append(Close[i])
            # init = True
            double_top = 0
            double_bottom = 0
            trend -= 1
            number = round(trend/3)
            neck_number = round(trend/5)
            mid_number = round(trend/2.5)
            percentage -= 0.0005
            # print(k)

            # trend = 300

            # number = 100
            # neck_number = 400
            # mid_number = 500
            # percentage = 2

            # if k != 0 and k < 3:
            #     neck_number -= 20   ### 360
            # elif k > 2 and k < 5:
            #     neck_number -= 20       ### 320
            #     mid_number -= 30        ### 440
            #     percentage -= 0.5       ### 1.5
            # elif k > 4 and k < 9:
            #     neck_number -= 20       ### 240
            #     mid_number -= 20        ### 360
            # elif k > 8 and k < 13:
            #     number -= 5        ### 80
            #     neck_number -= 20       ### 160
            #     mid_number -= 25        ### 260
            #     percentage -= 0.1       ### 1.1
            # elif k > 12 and k < 16:
            #     number -= 5         ### 65
            #     neck_number -= 20       ### 100
            #     mid_number -= 30        ### 170
            #     percentage -= 0.2       ### 0.5
            #     trend -= 25             ### 225
            # elif k > 15 and k < 18:
            #     number -= 5         ### 55
            #     neck_number -= 15   ### 70
            #     mid_number -= 30    ### 110
            #     percentage -= 0.1   ### 0.3
            # elif k > 17 and k < 20:
            #     number -= 10         ### 35
            #     neck_number -= 15    ### 40
            #     mid_number -= 30     ### 50
            #     percentage -= 0.1    ### 0.1
            #     trend -= 30         ### 165
            # elif k > 19 and k < 24:
            #     number -= 5        ###10
            #     neck_number -= 5   ### 20
            #     mid_number -= 5    ### 30
            #     percentage -= 0.018  ### 0.028
            #     trend -= 15         ### 105
            # elif k > 23 and k < 26:
            #     number -= 2         ### 6
            #     neck_number -= 5    ### 10
            #     mid_number -= 8     ### 14
            #     trend -= 35         ### 35
        print(trend)
        print(number)
        print(neck_number)
        print(mid_number)
        print(percentage)

        for k in range(1, 301):
            sma = []
            for i, (dte, l) in enumerate(zip(dframe['Date'], dframe['Low'])):
                if i >= trend:
                    sm = 0
                    j = trend
                    while j > 0:
                        sm += Close[i - j + 1]
                        j -= 1
                    try:
                        average = sm / trend
                        sma.append(average)
                        ###### الگو های بازگشتی و ادامه دهنده
                        ### روند صعودی
                        if sma[i] > sma[i - trend]:
                            #### double_top
                            if prev_price(number, i, Close) and double_top == 0 and next_price(neck_number, i,Close):  ### تشخیص سقف اول
                                double_top = 1
                                holder_i = i
                            elif (double_top == 1 and prev_price(neck_number, i, Close) and next_price(number, i, Close) and
                                Close[i] <= Close[holder_i] + Close[holder_i] * percentage and Close[i] >= Close[holder_i] - Close[
                                holder_i] * percentage):  ###سقف دوم اگر  فاصله اش با سقف قبلی کمتر از میدنامبر باشه و کمتر 3درصد سقف اول یا بیشتر 3 درصد سقف دوم باشه
                                double_top = 2
                                holder_i2 = i
                                volume_double_top = volume[i]
                                neck_price = Close[holder_i]
                                for j in range(holder_i, i):  ### پیدا کردن دره
                                    if Close[j] < neck_price:
                                        neck_price = Close[j]
                                        neck_i = j
                            elif double_top == 2 and Close[i] <= neck_price - neck_price * 0.03 and volume_double_top <= \
                                    volume[i]:  ### قطع گردن
                                double_top = 3
                            if double_top == 1 and i > holder_i + mid_number:
                                double_top = 0
                            if double_top == 2 and i > holder_i2 + mid_number:
                                double_top = 0
                            if double_top == 3:
                                double_top = 0
                                bifour = True
                                for m, n in zip(arr_holder_i, arr_double_top):
                                    if m <= i <= n:
                                        bifour = False
                                if bifour:
                                    arr_double_top.append(i)
                                    arr_holder_i.append(holder_i)
                                    fplt.add_text((dframe['Date'][holder_i], Close[holder_i]), f'---R---: {k}', color='red')
                                    fplt.add_text((dframe['Date'][holder_i2], Close[holder_i2]), f'---R2---: {k}',color='blue')
                                    fplt.add_text((dframe['Date'][neck_i], Close[neck_i]), f'---S---: {k}', color='green')
                                    target = Close[holder_i] - Close[neck_i]
                                    target -= Close[neck_i]
                                    fplt.add_text((dte, l - Low[i] / 40), f'-Doubletop: {target}')
                            ### double bottom
                            if double_bottom == 2 and Close[
                                i] >= bottom_neck_price + bottom_neck_price * 0.03 and volume_double_bottom <= volume[
                                i]:  ### قطع گردن
                                double_bottom = 3
                            if double_bottom == 1 and i > bottom_holder_i + mid_number:
                                double_bottom = 0
                            if double_bottom == 2 and i > bottom_holder_i2 + mid_number:
                                double_bottom = 0
                            if double_bottom == 3:
                                double_bottom = 0
                                bifour = True
                                for m, n in zip(bottom_arr_holder_i, bottom_arr_double_top):
                                    if m <= i <= n:
                                        bifour = False
                                if bifour:
                                    bottom_arr_double_top.append(i)
                                    bottom_arr_holder_i.append(bottom_holder_i)
                                    fplt.add_text((dframe['Date'][bottom_holder_i], Close[bottom_holder_i]),f'---R---: {k}', color='orange')
                                    fplt.add_text((dframe['Date'][bottom_holder_i2], Close[bottom_holder_i2]),f'---R2---: {k}', color='orange')
                                    fplt.add_text((dframe['Date'][bottom_neck_i], Close[bottom_neck_i]), f'---S---: {k}',color='orange')
                                    target = abs(Close[bottom_neck_i] - Close[bottom_holder_i])
                                    target += Close[bottom_neck_i]
                                    fplt.add_text((dte, l - Low[i] / 40), f'+Doublebottom: {target}')

                        ### روند نزولی
                        else:
                            ### double top
                            if double_top == 2 and Close[i] <= neck_price - neck_price * 0.03 and volume_double_top <= \
                                    volume[i]:  ### قطع گردن
                                double_top = 3
                            if double_top == 1 and i > holder_i + mid_number:
                                double_top = 0
                            if double_top == 2 and i > holder_i2 + mid_number:
                                double_top = 0
                            if double_top == 3:
                                double_top = 0
                                ### برای رسم سطوح حمایت و مقاومت
                                # resistance.append(three_holder_i)
                                # support.append(neck_i)
                                bifour = True
                                for m, n in zip(arr_holder_i, arr_double_top):
                                    if m <= i <= n:
                                        bifour = False
                                if bifour:
                                    arr_double_top.append(i)
                                    arr_holder_i.append(holder_i)
                                    fplt.add_text((dframe['Date'][holder_i], Close[holder_i]), f'---R---: {k}', color='red')
                                    fplt.add_text((dframe['Date'][holder_i2], Close[holder_i2]), f'---R2---: {k}',color='blue')
                                    fplt.add_text((dframe['Date'][neck_i], Close[neck_i]), f'---S---: {k}', color='green')
                                    target = Close[holder_i] - Close[neck_i]
                                    target -= Close[neck_i]
                                    fplt.add_text((dte, l - Low[i] / 40), f'-Doubletop: {target}')
                        ### double bottom
                        if prev_price(number, i, Close, False) and double_bottom == 0 and next_price(neck_number, i, Close,False):  ### تشخیص سقف اول
                            double_bottom = 1
                            bottom_holder_i = i
                        elif (double_bottom == 1 and prev_price(neck_number, i, Close, False) and next_price(number, i,Close,False) and
                            Close[i] <= Close[bottom_holder_i] + Close[bottom_holder_i] * percentage and Close[i] >=Close[bottom_holder_i] -Close[
                            bottom_holder_i] * percentage):  ###سقف دوم اگر  فاصله اش با سقف قبلی کمتر از میدنامبر باشه و کمتر 3درصد سقف اول یا بیشتر 3 درصد سقف دوم باشه
                            double_bottom = 2
                            bottom_holder_i2 = i
                            bottom_neck_price = Close[bottom_holder_i]
                            volume_double_bottom = volume[i]
                            for j in range(bottom_holder_i, i):  ### پیدا کردن دره
                                if Close[j] > bottom_neck_price:
                                    bottom_neck_price = Close[j]
                                    bottom_neck_i = j
                        elif double_bottom == 2 and Close[
                            i] >= bottom_neck_price + bottom_neck_price * 0.03 and volume_double_bottom <= volume[
                            i]:  ### قطع گردن
                            double_bottom = 3
                        if double_bottom == 1 and i > bottom_holder_i + mid_number:
                            double_bottom = 0
                        if double_bottom == 2 and i > bottom_holder_i2 + mid_number:
                            double_bottom = 0
                        if double_bottom == 3:
                            double_bottom = 0
                            bifour = True
                            for m, n in zip(bottom_arr_holder_i, bottom_arr_double_top):
                                if m <= i <= n:
                                    bifour = False
                            if bifour:
                                bottom_arr_double_top.append(i)
                                bottom_arr_holder_i.append(bottom_holder_i)
                                fplt.add_text((dframe['Date'][bottom_holder_i], Close[bottom_holder_i]), f'---R---: {k}',color='orange')
                                fplt.add_text((dframe['Date'][bottom_holder_i2], Close[bottom_holder_i2]), f'---R2---: {k}',color='orange')
                                fplt.add_text((dframe['Date'][bottom_neck_i], Close[bottom_neck_i]), f'---S---: {k}',color='orange')
                                target = abs(Close[bottom_neck_i] - Close[bottom_holder_i])
                                target += Close[bottom_neck_i]
                                fplt.add_text((dte, l - Low[i] / 40), f'+Doublebottom: {target}')
                                # fplt.add_text((dte, target), f'****: {k}')
                    except:
                        pass
                else:
                    sma.append(Close[i])
            double_top = 0
            double_bottom = 0
            trend -= 1
            number = round(trend / 3)
            neck_number = round(trend / 5)
            mid_number = round(trend / 2.5)
            percentage -= 0.0005


        ### رسم سطوح حمایت و مقاومت
        # for j in resistance:
        #     for dt in dframe.index:
        #         if Date[j] >= str(dt):
        #             fplt.add_text((dt, Close[j]), '---R---', color='red')
        # for j in support:
        #     for dt in dframe.index:
        #         if Date[j] >= str(dt):
        #             fplt.add_text((dt, Close[j]), '---S---', color='green')

        # rates_total = t_row_column(title, 'C')
        # wb = openpyxl.load_workbook(f'download_data/{title}.xlsx')
        # ws = wb.active
        # volume = []
        # volume_play = []
        # for t in range(rates_total):
        #     if t == 0:
        #         pass
        #     else:
        #         volume.append(ws[f'C{t + 1}'].value)
        #
        # # print(ws[f'C{rates_total}'].value)
        # return volume

    # for (int i=0;i < rates_total;i++)
    #     {
    #         Vol[i] = (double)(volume[i] * close[i]);
    #     }
    #     for (int i=Input1;i < rates_total;i++)
    #         {
    #             double
    #         Sum = 0;
    #         double
    #         Avarage = 0;
    #         for (int j=Input1; j > 0; j--)
    #         {
    #             Sum += Vol[i-j];
    #         }
    #         Avarage = Sum / Input1;
    #         prices[i] = ( Vol[i] / Avarage) * high[i];
    #         }
    else:
        pass



data = {
'Date': [
'2023-01-01' , '2023-01-02' , '2023-01-03' , '2023-01-04' , '2023-01-05' ,
'2023-01-06' , '2023-01-07' , '2023-01-08' , '2023-01-09' , '2023-01-10' ,
'2023-01-11' , '2023-01-12' , '2023-01-13' , '2023-01-14' , '2023-01-15' ,
'2023-01-16' , '2023-01-17' , '2023-01-18' , '2023-01-19' , '2023-01-20' , '2023-01-21' , '2023-01-22' , '2023-01-23' ,
'2023-01-24' , '2023-01-25' , '2023-01-26' , '2023-01-27' , '2023-01-28' ,
'2023-01-29' , '2023-01-30' , '2023-01-31'
]            ,

'open':  [1 , 2 , 3 , 5 ,  8 , 11 , 15 , 20 , 25 , 30 , 35 , 30 , 25 , 28 , 18 , 22 , 28 , 32 , 40 , 43 , 35 , 16 , 17 , 22 , 27 , 32 , 34 , 29 , 22 , 17 , 15] ,
'close': [2 , 3 , 5 , 8 , 11 , 15 , 20 , 25 , 30 , 35 , 30 , 25 , 28 , 17 , 22 , 28 , 32 , 40 , 44 , 34 , 25 , 18 , 22 , 28 , 33 , 36 , 29 , 21 , 18 ,  14 , 11] ,
'low':   [1 , 2 , 3 , 5 ,  7 , 10 , 14 , 19 , 24 , 28 , 32 , 28 , 24 , 27 , 26 , 16 , 11 ,  28 , 33 , 43 , 14 , 11 , 14 , 15 , 13 , 20 , 21 , 20 , 14 , 11 , 7] ,
'high':  [3 , 4 , 6 , 9 , 12 , 16 , 21 , 26 , 31 , 36 , 37 , 32 , 29 , 33 , 35 , 30 , 20 , 44 , 44 , 34 , 15 , 16 , 22 , 28 , 29 , 29 , 34 , 19 , 14 , 12 , 8]
}
# data = pd.read_excel('week-شستا.xlsx')
df = pd.DataFrame(data)
df['Date'] = pd.to_datetime(df['Date'])
df.set_index('Date', inplace=True)
Open = data['open']
High = data['high']
Low = data['low']
Close = data['close']
date = data['Date']
# Volume = data['volume']
# # ترسیم نمودار کندل استیک
ax = fplt.create_plot('نمودار قیمتی', rows=1)
candles = df[['open', 'close', 'high', 'low']]
fplt.candlestick_ochl(candles, ax=ax)
top_shift = {}
sma2 = []

percentage = 0.2

trend = 8
number = 4
neck_number = 4
mid_number = 10

three_top = 0
three_holder_i = 0
three_holder_i2 = 0
three_holder_i3 = 0
three_neck_price = 0
three_neck_price2 = 0
three_neck_i = 0
three_neck_i2 = 0
three_arr_holder_i = []
three_arr_top = []
three_volume_top = 0
init = False
for k in range(1, 2):
    sma2 = []
    for i, (dte, l) in enumerate(zip(df.index, df['low'])):
        if i >= trend:
            sm = 0
            j = trend
            while j > 0:
                sm += Close[i - j + 1]
                j -= 1
            average = sm / trend
            sma2.append(average)
            ###### الگو های بازگشتی و ادامه دهنده
            ### روند صعودی
            if sma2[i] > sma2[i - trend]:
                #### double_top
                # print(double_top)
                if prev_price(number, i, Close) and three_top == 0 and next_price(neck_number, i, Close):  ### تشخیص سقف اول
                    three_top = 1
                    print(Close[i])
                    # print(double_top)
                    three_holder_i = i
                elif (three_top == 1 and prev_price(neck_number, i, Close) and next_price(neck_number, i, Close) and
                      Close[i] >= Close[three_holder_i] - Close[three_holder_i] * percentage):
                    three_top = 2
                    three_holder_i2 = i
                    three_neck_price = Close[three_holder_i]
                    for j in range(three_holder_i, i):  ### پیدا کردن دره
                        if Close[j] < three_neck_price:
                            three_neck_price = Close[j]
                            three_neck_i = j
                    print('i2: ', Close[i])
                    print('neck: ', three_neck_price)
                elif (three_top == 2 and prev_price(neck_number, i, Close) and next_price(neck_number, i, Close) and
                      Close[i] >= Close[three_holder_i] - Close[three_holder_i] * percentage and Close[i] <= Close[three_holder_i] + Close[three_holder_i] * percentage ):
                    three_top = 3
                    three_holder_i3 = i
                    three_neck_price2 = Close[three_holder_i2]
                    for j in range(three_holder_i2, i):  ### پیدا کردن دره
                        if Close[j] < three_neck_price2:
                            three_neck_price2 = Close[j]
                            three_neck_i2 = j
                    print('i3: ', Close[i])
                    print('neck2: ', three_neck_price2)
                elif three_top == 2 and Close[i] <= three_neck_price - three_neck_price * 0.03:  ### قطع گردن
                    three_top = 3
                if three_top == 1 and i > three_holder_i + mid_number:
                    three_top = 0
                if three_top == 2 and i > three_holder_i2 + mid_number:
                    three_top = 0
                if three_top == 3:
                    three_top = 0
                    bifour = True
                    if init:
                        for m, n in zip(three_arr_holder_i, three_arr_top):
                            print('m:', m, "  n:", n,' i:',i)

                            if  (m <= i <= n):
                                bifour = False
                                print(bifour)
                                # print('m: ',m ,"     n: ",n)
                    if bifour:
                        three_arr_top.append(i)
                        three_arr_holder_i.append(three_holder_i)
                        fplt.add_text((df.index[three_holder_i], Close[three_holder_i]), f'---R---: {k}', color='red')
                        fplt.add_text((df.index[three_holder_i2], Close[three_holder_i2]), f'---R2---: {k}', color='blue')
                        fplt.add_text((df.index[three_neck_i], Close[three_neck_i]), f'---S---: {k}', color='green')
                        fplt.add_text((dte, l- Low[i] / 40 ), f'-Doubletop: {k}')

            ### روند نزولی
            else:

                if three_top == 2 and Close[i] <= three_neck_price - three_neck_price * 0.03:  ### قطع گردن
                    three_top = 3
                if three_top == 1 and i > three_holder_i + mid_number:
                    three_top = 0
                if three_top == 2 and i > three_holder_i2 + mid_number:
                    three_top = 0
                if three_top == 3:
                    three_top = 0
                    ### برای رسم سطوح حمایت و مقاومت
                    # resistance.append(three_holder_i)
                    # support.append(neck_i)
                    bifour = True
                    if init:
                        for m, n in zip(three_arr_holder_i, three_arr_top):
                            print('m:', m, "  n:", n,' i:',i)

                            if  (m <= i <= n):
                                bifour = False
                                print(bifour)

                    if bifour:
                        three_arr_top.append(i)
                        three_arr_holder_i.append(three_holder_i)
                        fplt.add_text((df.index[three_holder_i], Close[three_holder_i]), f'---R---: {k}', color='red')
                        fplt.add_text((df.index[three_holder_i2], Close[three_holder_i2]), f'---R2---: {k}', color='blue')
                        fplt.add_text((df.index[three_neck_i], Close[three_neck_i]), f'---S---: {k}', color='green')
                        fplt.add_text((dte, l - Low[i] / 40), f'-Doubletop: {k}')

        else:
            sma2.append(Close[i])
    init = True

    trend = 5
    number = 3
    neck_number = 2
    mid_number = 5
    percentage = 0.2
    three_top = 0
    # percentage -= 0.0005
    print('--------------')
fplt.plot(sma2, ax=ax)
fplt.show()
