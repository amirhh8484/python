#-----------------------------Amirhossein hasani--------------------------------
#---------------------------------bourse.py-------------------------------------
#-----download + check net + data convert + creat book + data append + save----- 

import csv
import os
import sys
import openpyxl
from openpyxl.workbook import Workbook
import pytse_client as tse
from pytse_client import download_client_types_records
import tkinter as tk
from tkinter import messagebox
import socket

current_directory = os.path.dirname(os.path.abspath(sys.executable))
folder_path = os.path.join(current_directory, 'tickers_data')
if not os.path.exists(folder_path):
    os.makedirs(folder_path)
folder_path2 = os.path.join(current_directory, 'client_types_data')
if not os.path.exists(folder_path2):
    os.makedirs(folder_path2)
folder_path3 = os.path.join(current_directory, 'tsetmc_adjust')
if not os.path.exists(folder_path3):
    os.makedirs(folder_path3)


def check_internet(host="8.8.8.8", port=53, timeout=3):
    try:
        socket.setdefaulttimeout(timeout)
        socket.socket(socket.AF_INET, socket.SOCK_STREAM).connect((host, port))
        return True
    except socket.error as ex:
        return False


def download1():
    if check_internet():
        tick = True
        client = True
        while tick or client:
            if tick:
                try:
                    tickers = tse.download(symbols="all", write_to_csv=True, include_jdate=True, adjust=True)
                    tick = False
                except Exception:
                    tick = True
            if client:
                try:
                    records_dict = download_client_types_records(symbols="all", write_to_csv=True, include_jdate=True)
                    client = False
                except Exception:
                    client = True
        messagebox.showinfo("دانلود دیتا", "دانلود با موفقیت انجام شد")
    else:
        messagebox.showinfo("عدم اتصال به اینترنت", "خطا در اتصال به اینترنت ")

def is_file_in_folders(file_name):
    if file_name != '18719101.csv':
        file_in_folder1 = os.path.isfile(os.path.join('tickers_data', file_name))
        ln = len(file_name)
        file_name = file_name[:ln-6] + file_name[ln-4:]
        file_in_folder2 = os.path.isfile(os.path.join('client_types_data', file_name))
        return file_in_folder1 and file_in_folder2


def sht_val(x, smbl):
    wd = Workbook()
    sheet3 = wd.active
    with open(f'client_types_data//{smbl}.csv', mode='r', newline='', encoding='utf-8') as csv_file:
        csv_reader = csv.reader(csv_file)
        for i, row in enumerate(csv_reader):
            if i >= x:
                break
            new_row = []
            for cell in row:
                try:
                    new_row.append(int(cell))
                except ValueError:
                    try:
                        new_row.append(float(cell))
                    except ValueError:
                        new_row.append(cell)
        sheet3.append(new_row)
    wd.save('mc.xlsx')


def convert_csv_to_xlsx():
    files = []
    for item in os.listdir('tickers_data'):
        item_path = os.path.join('tickers_data', item)
        if os.path.isfile(item_path):
            if is_file_in_folders(item):
                files.append(item)
    col_ticker = ["A", "C", "D", "E", "F", "G", "H"]
    col_t_client = ["B", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y"]
    col_client = ["T", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "Q", "P", "R", "S"]
    for file in files:
        smbl = file[:-6]
        if not os.path.isfile(f'tsetmc_adjust/{smbl}.xlsx'):
            wm = Workbook()
            wm.save(f'tsetmc_adjust/{smbl}.xlsx')
        wt = openpyxl.load_workbook(filename=f'tsetmc_adjust/{smbl}.xlsx')
        wz2 = Workbook()
        ws2 = wz2.active
        with open(f'tickers_data//{file[:-4]}.csv', mode='r', newline='', encoding='utf-8') as csv_file2:
            csv_reader2 = csv.reader(csv_file2)
            for row2 in csv_reader2:
                new_row2 = []
                for cell2 in row2:
                    try:
                        new_row2.append(int(cell2))
                    except ValueError:
                        try:
                            new_row2.append(float(cell2))
                        except ValueError:
                            new_row2.append(cell2)
                ws2.append(new_row2)
        wz2.save('mc2.xlsx')
        wd = openpyxl.load_workbook(filename='mc2.xlsx')
        sht_tic = wd["Sheet"]
        if smbl in wt.sheetnames:
            sht_tset = wt[smbl]
            tf = True
            x = 2
            count = 0
            while tf:
                sht_val(x, smbl)
                wb = openpyxl.load_workbook(filename='mc.xlsx')
                sht_client = wb["Sheet"]
                date_client = sht_client['T1'].value
                date_tset = sht_tset[f'B{x}'].value
                if date_client is not None and date_tset is not None:
                    if sht_tset[f'B{x}'].value == date_client:
                        tf = False
                    if date_client > date_tset:
                        sht_tset.insert_rows(x)
                        for itm, itm2 in zip(col_t_client, col_client):
                            sht_tset[f"{itm}{x}"].value = sht_client[f"{itm2}1"].value
                        for i in col_ticker:
                            sht_tset[f"{i}{x}"].value = sht_tic[f"{i}{sht_tic.max_row - count}"].value
                x += 1
                count += 1
        else:
            wz = Workbook()
            ws = wz.active
            with open(f'client_types_data//{smbl}.csv', mode='r', newline='', encoding='utf-8') as csv_file:
                csv_reader = csv.reader(csv_file)
                for row in csv_reader:
                    new_row = []
                    for cell in row:
                        try:
                            new_row.append(int(cell))
                        except ValueError:
                            try:
                                new_row.append(float(cell))
                            except ValueError:
                                new_row.append(cell)
                    ws.append(new_row)
            wz.save('mc.xlsx')
            wb = openpyxl.load_workbook(filename='mc.xlsx')
            sht_client = wb["Sheet"]
            wt.create_sheet(title=smbl)
            sht_tset = wt[smbl]
            for i in col_ticker:
                bl = True
                for k, k2 in zip(range(1, sht_tic.max_row + 1), range(sht_tic.max_row + 1, 1, -1)):
                    if bl:
                        sht_tset[f"{i}{k}"].value = sht_tic[f"{i}{k}"].value
                        bl = False
                    else:
                        sht_tset[f"{i}{k}"].value = sht_tic[f"{i}{k2}"].value
            for item3, item2 in zip(col_t_client, col_client):
                for k in range(1, sht_client.max_row + 1):
                    sht_tset[f"{item3}{k}"].value = sht_client[f"{item2}{k}"].value
        wt.save(f'tsetmc_adjust/{smbl}.xlsx')
    file_name = 'mc.xlsx'
    file_name2 = 'mc2.xlsx'
    file_path = os.path.join(os.path.dirname(os.path.abspath(sys.executable)), file_name)
    file_path2 = os.path.join(os.path.dirname(os.path.abspath(sys.executable)), file_name2)
    if os.path.exists(file_path):
        os.remove(file_path)
    if os.path.exists(file_path2):
        os.remove(file_path2)
    lbl9.config(text='عملیات با موفقیت به پایان رسید', font=('Arial', 12), fg='green')


root = tk.Tk()
root.title("DB-Bourse-Adjust-Separate")
root.geometry('350x350')
root.minsize(300, 300)
root.maxsize(400, 400)
lbl = tk.Label(root, text='')
lbl.pack()
lbl2 = tk.Label(root, text='')
lbl2.pack()
lbl3 = tk.Label(root, text='')
lbl3.pack()
btn = tk.Button(root, text='   Data Download   ', command=download1, font=('Arial', 14))
btn.pack()
lbl4 = tk.Label(root, text='')
lbl4.pack()
lbl5 = tk.Label(root, text='')
lbl5.pack()
lbl6 = tk.Label(root, text='')
lbl6.pack()
btn3 = tk.Button(root, text='   Excel Output   ', command=convert_csv_to_xlsx, font=('Arial', 14))
btn3.pack()
lbl8 = tk.Label(root, text='')
lbl8.pack()
lbl9 = tk.Label(root, text='')
lbl9.pack()
root.mainloop()
